"""존(Zone) 분석 — 원판(disk) 마스크 집합 차집합 기반 존 분할·퍼센티지 계산.

스펙: docs/specs/zone-analysis-tab-2026-08-25.md "판단 2"·"존 계산 로직" 절.
Qt 의존성 없음(core 규칙) — 순수 numpy 함수.

원을 반지름 오름차순 [C_0(최소), ..., C_{N-1}(최대)]로 정렬했을 때:
    Zone_center = mask(C_0)
    Zone_i      = mask(C_{i+1}) AND NOT mask(C_i)   (i = 0..N-2)
    Zone_outside = 전체 이미지 AND NOT mask(C_{N-1})
전제: 원들이 대략 nested(중첩)돼 있다 — 배터리 캡 실제 구조상 항상 성립. 수동 편집으로
원을 교차시키는 비정상 입력은 v1에서 방지하지 않는다(스펙 명시, YAGNI).
"""
from dataclasses import dataclass
from pathlib import Path

import cv2
import numpy as np


@dataclass
class Circle:
    id: int
    cx: float
    cy: float
    r: float


@dataclass
class Zone:
    index: int          # 0=중심부, 1..N-1=링, N=바깥쪽 (원 N개 기준)
    name: str
    mask: np.ndarray     # (H, W) bool


def disk_mask(cx: float, cy: float, r: float, img_shape: tuple[int, int]) -> np.ndarray:
    """(x-cx)^2 + (y-cy)^2 <= r^2 벡터화 비교로 원판 마스크 생성 — fillPoly 아님.

    R3-2에서 공개 전환(언더스코어 제거) — `zone_canvas.py`의 브러시 지우기 undo
    복원(전체 재생)이 이 함수를 그대로 재사용한다(스펙 판단 2).
    """
    h, w = img_shape
    yy, xx = np.ogrid[:h, :w]
    return (xx - cx) ** 2 + (yy - cy) ** 2 <= r ** 2


def zones_from_circles(circles: list[Circle], img_shape: tuple[int, int]) -> list[Zone]:
    """반지름 오름차순 정렬 후 원판 마스크 차집합으로 존 목록 생성.

    원이 없으면 빈 리스트(존 개념 자체가 성립하지 않음). 원 N개면 존 N+1개.
    """
    if not circles:
        return []
    sorted_c = sorted(circles, key=lambda c: c.r)
    n = len(sorted_c)
    masks = [disk_mask(c.cx, c.cy, c.r, img_shape) for c in sorted_c]

    zones = [Zone(0, "중심부", masks[0])]
    for i in range(n - 1):
        zones.append(Zone(i + 1, f"링 {i + 1}", masks[i + 1] & ~masks[i]))
    zones.append(Zone(n, "바깥쪽", ~masks[-1]))
    return zones


def apply_manual_strokes(
    mask: np.ndarray,
    manual_strokes: list[tuple[bool, list[tuple[float, float, float]]]],
) -> np.ndarray:
    """수동 그리기/지우기를 시간순으로 적용(마지막 스트로크 우선). `disk_mask()` 재사용.

    R-ZONE-3에서 `ZoneCanvas.apply_manual_strokes()`(인스턴스 메서드)를 순수
    함수로 승격 — 현재 화면 캔버스가 아닌 다른 이미지의 사이드카 상태(배치
    처리 중 재조회한 `manual_strokes`)에도 적용할 수 있어야 하기 때문이다.
    """
    result = mask.copy()
    for draw, stroke in manual_strokes:
        stroke_mask = np.zeros(mask.shape, dtype=bool)
        for cx, cy, r in stroke:
            stroke_mask |= disk_mask(cx, cy, r, mask.shape)
        result[stroke_mask] = draw
    return result


def zone_stats(zone_mask: np.ndarray, target_class_mask: np.ndarray) -> float:
    """존 마스크 면적 대비 타겟 클래스(AND) 픽셀 비율(%)."""
    area = int(zone_mask.sum())
    if area == 0:
        return 0.0
    hit = int((zone_mask & target_class_mask).sum())
    return hit / area * 100.0


def compute_blob_labels(mask: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """타겟 클래스 이진 마스크 -> (라벨맵, stats, centroids). 블랍 클릭 삭제 +
    R3 Excel blob 내보내기가 공유하는 헬퍼.

    상하좌우로 이어진 픽셀만 같은 블랍으로 보는 4-connectivity를 사용한다 —
    `inference_engine._compute_blobs_and_filter()`와 API는 비슷하지만 confidence/
    size threshold 필터링까지 가져오면 이 탭엔 불필요한 결합이 생겨(스펙 "블랍 삭제"
    절) 별도로 둔다. 라벨 0 = 배경. `stats[label] = [x, y, w, h, area]`(OpenCV 표준
    컬럼 순서, `CC_STAT_LEFT/TOP/WIDTH/HEIGHT/AREA`). `centroids[label] = (cx, cy)`
    (R3 — `zone_blob_stats()`가 zone 배정에 사용, 이미 계산되는 값이라 추가 비용 없음).
    """
    _, labels, stats, centroids = cv2.connectedComponentsWithStats(
        mask.astype(np.uint8), connectivity=4
    )
    return labels, stats, centroids


@dataclass
class ZoneBlobStat:
    """R3 — zone별 blob 하나의 통계(Excel `zone_blobs` 시트 1행에 대응)."""
    zone_name: str
    blob_id: int             # final_mask 라벨 순서 기준 1부터
    pixel_count: int
    ai_score: float | None   # 0~1, 순수 수동 blob(ai_mask 교집합 없음)이면 None
    centroid_x: float
    centroid_y: float
    bbox_x: int
    bbox_y: int
    bbox_w: int
    bbox_h: int


def zone_blob_stats(
    zones: list[Zone],
    ai_mask: np.ndarray,
    final_mask: np.ndarray,
    confidence_map: np.ndarray,
) -> list[ZoneBlobStat]:
    """blob(연결요소, `final_mask` 기준) 단위로 zone 배정 + AI 점수를 계산한다.

    blob의 "모양"은 최종 편집 결과(`final_mask`, AI 예측 + 수동 그리기/지우기 반영)
    기준으로 나누되, AI 점수는 원래 모델이 예측한 픽셀(`ai_mask`)에서만 confidence
    평균을 낸다 — 사람이 브러시로 새로 그려 넣은 픽셀은 confidence 자체가 없으므로
    자동으로 제외되고, 그 blob 전체의 AI 예측 부분 평균이 대표값으로 쓰인다(스펙
    "R3" 판단 2). 교집합이 전혀 없으면(100% 수동 blob) `ai_score=None`.
    zone 배정은 blob 중심점(centroid) 기준(스펙 판단 3).
    """
    labels, stats, centroids = compute_blob_labels(final_mask)
    n_labels = stats.shape[0]

    flat_labels = labels.ravel()
    ai_flat = ai_mask.ravel()
    conf_flat = confidence_map.ravel().astype(np.float64)
    ai_labels = flat_labels[ai_flat]
    ai_conf = conf_flat[ai_flat]
    ai_cnt = np.bincount(ai_labels, minlength=n_labels)
    ai_sum = np.bincount(ai_labels, weights=ai_conf, minlength=n_labels)

    h, w = final_mask.shape
    results: list[ZoneBlobStat] = []
    for lbl in range(1, n_labels):   # 0 = 배경
        pixel_count = int(stats[lbl, cv2.CC_STAT_AREA])
        if pixel_count == 0:
            continue
        cnt = int(ai_cnt[lbl])
        ai_score = float(ai_sum[lbl] / cnt) if cnt > 0 else None

        cx, cy = centroids[lbl]
        px = min(max(int(round(cx)), 0), w - 1)
        py = min(max(int(round(cy)), 0), h - 1)
        zone_name = "미분류"   # 존들이 전체 이미지를 파티션하므로 실제로는 발생하지 않음
        for zone in zones:
            if zone.mask[py, px]:
                zone_name = zone.name
                break

        results.append(ZoneBlobStat(
            zone_name=zone_name,
            blob_id=lbl,
            pixel_count=pixel_count,
            ai_score=ai_score,
            centroid_x=float(cx),
            centroid_y=float(cy),
            bbox_x=int(stats[lbl, cv2.CC_STAT_LEFT]),
            bbox_y=int(stats[lbl, cv2.CC_STAT_TOP]),
            bbox_w=int(stats[lbl, cv2.CC_STAT_WIDTH]),
            bbox_h=int(stats[lbl, cv2.CC_STAT_HEIGHT]),
        ))
    return results


def _zone_name_sort_key(name: str) -> tuple[int, int]:
    """중심부 -> 링 N(N 오름차순) -> 바깥쪽 순 정렬 키 (스펙 R3-2 판단 4)."""
    if name == "중심부":
        return (0, 0)
    if name == "바깥쪽":
        return (2, 0)
    import re
    m = re.search(r"\d+", name)
    return (1, int(m.group()) if m else 0)


def pivot_wide_format(
    rows: list[tuple[str, str, float]]
) -> tuple[list[str], list[str], dict[tuple[str, str], float]]:
    """(이미지파일명, 존이름, 퍼센티지) long rows -> (이미지목록, 정렬된 존이름 열목록, 값 dict).

    없는 (이미지, 존) 조합은 값 dict에 키가 없음 -> 호출부가 공란/N-A로 렌더링.
    이미지 목록은 rows에 처음 등장한 순서 유지, 존 이름 열은 전체 rows에서 관측된
    합집합을 중심부->링 N(오름차순)->바깥쪽 순으로 정렬(스펙 판단 4, 원 개수가
    이미지마다 달라도 자연스러운 중심->외곽 순서 유지).
    """
    images: list[str] = []
    seen_images: set[str] = set()
    zone_names: set[str] = set()
    values: dict[tuple[str, str], float] = {}
    for image_name, zone_name, pct in rows:
        if image_name not in seen_images:
            seen_images.add(image_name)
            images.append(image_name)
        zone_names.add(zone_name)
        values[(image_name, zone_name)] = pct
    zone_cols = sorted(zone_names, key=_zone_name_sort_key)
    return images, zone_cols, values


def export_zone_percentages_to_excel(
    rows: list[tuple[str, str, float]],
    out_path: Path,
    blob_rows: list[tuple[str, ZoneBlobStat]] | None = None,
) -> None:
    """(이미지파일명, 존이름, 타겟비율%) long format 목록을 xlsx로 저장.

    `inference_engine.export_blobs_to_excel()`과 동일한 openpyxl 패턴(헤더만 볼드)을
    스키마만 바꿔 복제. R3-2부터 시트 2개 — 기존 long 시트("zones") 유지 +
    `pivot_wide_format()` 기반 wide 시트("zones_wide") 추가(애디티브, 시그니처 불변
    — 기존 호출부 전부 무변경으로 wide 시트를 자동으로 얻는다).

    R3: `blob_rows`가 주어지면(선택 인자, 하위 호환) 3번째 시트 "zone_blobs" 추가 —
    zone별 blob 크기 + AI 점수(`ZoneBlobStat.ai_score is None`이면 "N/A").
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "zones"
    ws.append(["이미지파일명", "존이름", "타겟비율(%)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for image_name, zone_name, pct in rows:
        ws.append([image_name, zone_name, round(pct, 2)])

    images, zone_cols, values = pivot_wide_format(rows)
    ws2 = wb.create_sheet("zones_wide")
    ws2.append(["이미지파일명"] + zone_cols)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for img in images:
        ws2.append(
            [img] + [round(values[(img, z)], 2) if (img, z) in values else "" for z in zone_cols]
        )

    if blob_rows:
        ws3 = wb.create_sheet("zone_blobs")
        ws3.append([
            "이미지파일명", "존이름", "blob_id", "픽셀수(면적)", "AI 점수(%)",
            "중심x", "중심y", "bbox_x", "bbox_y", "bbox_w", "bbox_h",
        ])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
        for image_name, s in blob_rows:
            ai_score = "N/A" if s.ai_score is None else round(s.ai_score * 100, 2)
            ws3.append([
                image_name, s.zone_name, s.blob_id, s.pixel_count, ai_score,
                round(s.centroid_x, 1), round(s.centroid_y, 1),
                s.bbox_x, s.bbox_y, s.bbox_w, s.bbox_h,
            ])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


if __name__ == "__main__":
    # ── self-check: 5x5 합성 이미지, 원 1개(cx=2, cy=2, r=1) — 손계산 검산 ──────
    # (x-2)^2+(y-2)^2<=1 만족 픽셀: (2,1)(1,2)(2,2)(3,2)(2,3) = 5개 (전체 25개 중)
    shape = (5, 5)
    circles = [Circle(1, 2, 2, 1)]
    zones = zones_from_circles(circles, shape)
    assert len(zones) == 2, "원 1개 -> 존 2개(중심부/바깥쪽)"
    assert zones[0].name == "중심부" and zones[1].name == "바깥쪽"
    assert zones[0].mask.sum() == 5, f"중심부 면적 5 예상, 실측 {zones[0].mask.sum()}"
    assert zones[1].mask.sum() == 20, f"바깥쪽 면적 20 예상, 실측 {zones[1].mask.sum()}"

    # 타겟 마스크 = 중심부와 정확히 일치 -> 중심부 100%, 바깥쪽 0%
    target = zones[0].mask
    assert zone_stats(zones[0].mask, target) == 100.0
    assert zone_stats(zones[1].mask, target) == 0.0

    # ── 원 2개 중첩 — 파티션 불변식: 존들의 합집합 = 전체 이미지, 겹침 없음 ──────
    circles2 = [Circle(1, 4, 4, 1), Circle(2, 4, 4, 3)]
    zones2 = zones_from_circles(circles2, shape)
    assert len(zones2) == 3
    assert [z.name for z in zones2] == ["중심부", "링 1", "바깥쪽"]
    total = sum(int(z.mask.sum()) for z in zones2)
    assert total == shape[0] * shape[1], "존 면적 합 = 전체 픽셀 수(겹침/누락 없음)"
    for a in zones2:
        for b in zones2:
            if a is not b:
                assert not (a.mask & b.mask).any(), "존끼리 겹치면 안 됨"

    # 타겟 마스크 = 전체 이미지(전부 녹) -> 모든 존 100%
    all_ones = np.ones(shape, dtype=bool)
    for z in zones2:
        assert zone_stats(z.mask, all_ones) == 100.0

    # 원 0개 -> 존 없음
    assert zones_from_circles([], shape) == []

    # ── compute_blob_labels: 서로 떨어진 블랍 2개 검산 ──────────────────────
    blob_mask = np.zeros((6, 6), dtype=np.uint8)
    blob_mask[0:2, 0:2] = 1   # 블랍 A, 면적 4
    blob_mask[4:6, 4:6] = 1   # 블랍 B, 면적 4
    labels, stats, centroids = compute_blob_labels(blob_mask)
    assert labels.shape == blob_mask.shape
    unique_labels = set(np.unique(labels).tolist()) - {0}
    assert len(unique_labels) == 2, f"블랍 2개 예상, 실측 {len(unique_labels)}"
    for lbl in unique_labels:
        assert stats[lbl, 4] == 4, f"블랍 면적 4 예상, 실측 {stats[lbl, 4]}"
    assert labels[0, 0] != labels[5, 5] and labels[0, 0] != 0 and labels[5, 5] != 0
    assert centroids.shape == (3, 2), "centroids: 배경(0) 포함 3행"

    # ── pivot_wide_format: 이미지 3장, 존 개수 2/2/3개 섞은 케이스(R3-2) ────────
    wide_rows = [
        ("a.jpg", "중심부", 10.0), ("a.jpg", "바깥쪽", 20.0),
        ("b.jpg", "중심부", 30.0), ("b.jpg", "바깥쪽", 40.0),
        ("c.jpg", "중심부", 50.0), ("c.jpg", "링 1", 60.0), ("c.jpg", "바깥쪽", 70.0),
    ]
    images, zone_cols, values = pivot_wide_format(wide_rows)
    assert images == ["a.jpg", "b.jpg", "c.jpg"], f"이미지 순서 유지 실패: {images}"
    assert zone_cols == ["중심부", "링 1", "바깥쪽"], f"정렬 순서(중심부->링->바깥쪽) 실패: {zone_cols}"
    assert values[("a.jpg", "중심부")] == 10.0
    assert ("a.jpg", "링 1") not in values, "없는 조합은 값 dict에 키가 없어야 함(공란 렌더링용)"
    assert values[("c.jpg", "링 1")] == 60.0

    # ── apply_manual_strokes: 마지막 스트로크가 우선(last-write-wins) ──────────
    base = np.zeros((9, 9), dtype=bool)
    strokes = [(True, [(4.0, 4.0, 2.0)]), (False, [(4.0, 4.0, 1.0)])]
    edited = apply_manual_strokes(base, strokes)
    assert edited[4, 2], "그리기 스트로크 반경 밖(x=2)은 True로 남아야 함"
    assert not edited[4, 4], "지우기 스트로크가 마지막이라 중심(4,4)은 False여야 함"

    # ── zone_blob_stats: AI blob 1개(교집합 있음) + 순수 수동 blob 1개(교집합
    # 없음) + zone 2개(중심부/바깥쪽)에 걸친 배정 확인(R3) ──────────────────────
    shape2 = (10, 10)
    zones3 = zones_from_circles([Circle(1, 2, 2, 1)], shape2)   # 중심부(2,2 부근)/바깥쪽
    final2 = np.zeros(shape2, dtype=bool)
    final2[2, 2] = True                      # blob A: 중심부 zone, AI 예측 픽셀
    final2[8, 8] = True                      # blob B: 바깥쪽 zone, 순수 수동(AI 교집합 없음)
    ai_mask2 = np.zeros(shape2, dtype=bool)
    ai_mask2[2, 2] = True                    # blob A만 AI가 실제로 예측
    conf_map2 = np.zeros(shape2, dtype=np.float32)
    conf_map2[2, 2] = 0.75

    blob_stats = zone_blob_stats(zones3, ai_mask2, final2, conf_map2)
    assert len(blob_stats) == 2, f"blob 2개 예상, 실측 {len(blob_stats)}"
    by_pos = {(s.centroid_x, s.centroid_y): s for s in blob_stats}
    blob_a = by_pos[(2.0, 2.0)]
    blob_b = by_pos[(8.0, 8.0)]
    assert blob_a.zone_name == "중심부", f"blob A는 중심부 배정 예상, 실측 {blob_a.zone_name}"
    assert blob_a.ai_score is not None and abs(blob_a.ai_score - 0.75) < 1e-6, \
        f"blob A AI 점수 0.75 예상, 실측 {blob_a.ai_score}"
    assert blob_b.zone_name == "바깥쪽", f"blob B는 바깥쪽 배정 예상, 실측 {blob_b.zone_name}"
    assert blob_b.ai_score is None, f"순수 수동 blob은 AI 점수 None(N/A) 예상, 실측 {blob_b.ai_score}"
    assert blob_a.pixel_count == 1 and blob_b.pixel_count == 1

    print("zone_metrics self-check OK")
